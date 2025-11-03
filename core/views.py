from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from orders.models import Order, Service
from partners.models import MitraProfile, Laundry, MitraRequest
from math import radians, sin, cos, sqrt, atan2
from django.http import HttpResponse
from datetime import datetime, timedelta
from django.db.models import Sum, Count, Q
import csv
import json

def home(request):
    if request.user.is_authenticated:
        return redirect('core:dashboard')
    
    services = Service.objects.filter(is_active=True)[:3]
    context = {
        'services': services
    }
    return render(request, 'core/home.html', context)

@login_required
def dashboard(request):
    user = request.user
    
    if user.role == 'admin':
        return redirect('core:admin_dashboard')
    elif user.role == 'mitra':
        return redirect('core:mitra_dashboard')
    else:
        # Get all active laundries in Yogyakarta
        laundries = Laundry.objects.filter(is_active=True, city='Yogyakarta').select_related('mitra')
        
        # Calculate user order statistics
        user_orders = Order.objects.filter(user=user)
        total_orders = user_orders.count()
        active_orders = user_orders.filter(status__in=['pending', 'picked_up', 'processing', 'ready']).count()
        nearby_laundries_count = laundries.count()
        
        # Enrich laundries with additional info
        for laundry in laundries:
            # You can add distance calculation here if user location is available
            # For now, just add a default distance
            laundry.distance = None
        
        context = {
            'laundries': laundries,
            'total_orders': total_orders,
            'active_orders': active_orders,
            'nearby_laundries_count': nearby_laundries_count,
        }
        return render(request, 'core/user_dashboard_new.html', context)

@login_required
def admin_dashboard(request):
    if request.user.role != 'admin':
        return redirect('core:dashboard')
    
    # Get basic statistics
    total_orders = Order.objects.count()
    total_users = request.user.__class__.objects.filter(role='user').count()
    total_mitras = Laundry.objects.filter(is_active=True).count()
    
    # Calculate revenue
    delivered_orders = Order.objects.filter(status='delivered')
    total_revenue = delivered_orders.aggregate(Sum('total_price'))['total_price__sum'] or 0
    platform_fees = delivered_orders.aggregate(Sum('platform_fee'))['platform_fee__sum'] or 0
    
    # New users this month
    now = datetime.now()
    start_of_month = datetime(now.year, now.month, 1)
    new_users_this_month = request.user.__class__.objects.filter(
        role='user',
        date_joined__gte=start_of_month
    ).count()
    
    # Orders this week
    start_of_week = now - timedelta(days=now.weekday())
    orders_this_week = Order.objects.filter(created_at__gte=start_of_week).count()
    
    # Pending mitra requests
    mitra_requests = MitraRequest.objects.filter(status='pending').select_related('user')
    pending_requests = mitra_requests.count()
    
    # Revenue trend (last 7 days)
    revenue_data = []
    revenue_labels = []
    for i in range(6, -1, -1):
        date = now - timedelta(days=i)
        daily_revenue = Order.objects.filter(
            status='delivered',
            created_at__date=date.date()
        ).aggregate(Sum('total_price'))['total_price__sum'] or 0
        revenue_data.append(float(daily_revenue))
        revenue_labels.append(date.strftime('%d %b'))
    
    # Status counts for pie chart
    status_counts = [
        Order.objects.filter(status='pending').count(),
        Order.objects.filter(status='picked_up').count(),
        Order.objects.filter(status='processing').count(),
        Order.objects.filter(status='ready').count(),
        Order.objects.filter(status='delivered').count(),
        Order.objects.filter(status='cancelled').count(),
    ]
    
    # Recent orders
    recent_orders = Order.objects.all().select_related('user', 'laundry').order_by('-created_at')[:50]
    
    context = {
        'total_orders': total_orders,
        'total_users': total_users,
        'total_mitras': total_mitras,
        'total_revenue': total_revenue,
        'platform_fees': platform_fees,
        'new_users_this_month': new_users_this_month,
        'orders_this_week': orders_this_week,
        'pending_requests': pending_requests,
        'mitra_requests': mitra_requests[:5],  # Show only first 5
        'revenue_data': json.dumps(revenue_data),
        'revenue_labels': json.dumps(revenue_labels),
        'status_counts': json.dumps(status_counts),
        'recent_orders': recent_orders,
    }
    return render(request, 'core/admin_dashboard_new.html', context)

@login_required
def mitra_dashboard(request):
    if request.user.role != 'mitra':
        return redirect('core:dashboard')
    
    try:
        mitra_profile = request.user.mitra_profile
    except:
        # Mitra profile doesn't exist yet
        messages.info(request, 'Selamat! Akun Anda telah disetujui sebagai mitra. Silakan lengkapi profil laundry Anda terlebih dahulu.')
        # For now, show empty dashboard or redirect to setup
        context = {
            'laundry': None,
            'orders': [],
            'pending_orders': 0,
            'processing_orders': 0,
            'completed_orders': 0,
            'total_revenue': 0,
            'needs_setup': True
        }
        return render(request, 'core/mitra_dashboard_new.html', context)
    
    # Get mitra's laundry
    laundry = Laundry.objects.filter(mitra=mitra_profile, is_active=True).first()
    
    if not laundry:
        # If no laundry yet, show empty dashboard with setup message
        messages.info(request, 'Silakan daftarkan usaha laundry Anda untuk mulai menerima pesanan.')
        context = {
            'laundry': None,
            'orders': [],
            'pending_orders': 0,
            'processing_orders': 0,
            'completed_orders': 0,
            'total_revenue': 0,
            'needs_setup': True
        }
        return render(request, 'core/mitra_dashboard_new.html', context)
    
    # Get all orders for this laundry
    all_orders = Order.objects.filter(laundry=laundry).select_related('user').order_by('-created_at')
    
    # Calculate statistics
    pending_orders = all_orders.filter(status='pending').count()
    processing_orders = all_orders.filter(status__in=['picked_up', 'processing']).count()
    completed_orders = all_orders.filter(status='delivered').count()
    
    # Calculate total revenue from delivered orders
    total_revenue = sum(
        order.laundry_price for order in all_orders.filter(status='delivered')
    )
    
    # Get recent orders (last 20)
    orders = all_orders[:20]
    
    context = {
        'laundry': laundry,
        'orders': orders,
        'pending_orders': pending_orders,
        'processing_orders': processing_orders,
        'completed_orders': completed_orders,
        'total_revenue': total_revenue,
        'needs_setup': False
    }
    return render(request, 'core/mitra_dashboard_new.html', context)

@login_required
def export_orders_excel(request):
    """Export all orders to Excel format"""
    if request.user.role != 'admin':
        return redirect('core:dashboard')
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders"
        
        # Headers
        headers = [
            'Order Number', 'Date', 'User', 'Email', 'Laundry', 'District',
            'Weight (kg)', 'Distance (km)', 'Laundry Price', 'COD Fee',
            'Platform Fee', 'Total Price', 'Payment Method', 'Status'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        orders = Order.objects.all().select_related('user', 'laundry').order_by('-created_at')
        
        for row, order in enumerate(orders, 2):
            ws.cell(row=row, column=1, value=order.order_number)
            ws.cell(row=row, column=2, value=order.created_at.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=3, value=order.user.get_full_name())
            ws.cell(row=row, column=4, value=order.user.email)
            ws.cell(row=row, column=5, value=order.laundry.name if order.laundry else '-')
            ws.cell(row=row, column=6, value=order.laundry.district if order.laundry else '-')
            ws.cell(row=row, column=7, value=float(order.weight_kg))
            ws.cell(row=row, column=8, value=float(order.distance_km) if order.distance_km else 0)
            ws.cell(row=row, column=9, value=float(order.laundry_price) if order.laundry_price else 0)
            ws.cell(row=row, column=10, value=float(order.cod_fee) if order.cod_fee else 0)
            ws.cell(row=row, column=11, value=float(order.platform_fee) if order.platform_fee else 0)
            ws.cell(row=row, column=12, value=float(order.total_price))
            ws.cell(row=row, column=13, value=order.get_payment_method_display())
            ws.cell(row=row, column=14, value=order.get_status_display())
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        return response
        
    except ImportError:
        # If openpyxl not installed, return CSV instead
        return export_orders_csv(request)

@login_required
def export_orders_csv(request):
    """Export all orders to CSV format"""
    if request.user.role != 'admin':
        return redirect('core:dashboard')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename=orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    
    writer = csv.writer(response)
    
    # Headers
    writer.writerow([
        'Order Number', 'Date', 'User', 'Email', 'Laundry', 'District',
        'Weight (kg)', 'Distance (km)', 'Laundry Price', 'COD Fee',
        'Platform Fee', 'Total Price', 'Payment Method', 'Status'
    ])
    
    # Data
    orders = Order.objects.all().select_related('user', 'laundry').order_by('-created_at')
    
    for order in orders:
        writer.writerow([
            order.order_number,
            order.created_at.strftime('%Y-%m-%d %H:%M'),
            order.user.get_full_name(),
            order.user.email,
            order.laundry.name if order.laundry else '-',
            order.laundry.district if order.laundry else '-',
            float(order.weight_kg),
            float(order.distance_km) if order.distance_km else 0,
            float(order.laundry_price) if order.laundry_price else 0,
            float(order.cod_fee) if order.cod_fee else 0,
            float(order.platform_fee) if order.platform_fee else 0,
            float(order.total_price),
            order.get_payment_method_display(),
            order.get_status_display(),
        ])
    
    return response

